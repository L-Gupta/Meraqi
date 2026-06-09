"""Automated multi-tab Excel databook generator."""

import json
import logging
from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.storage import deal_store, file_store

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


class DatabookError(Exception):
    pass


def _load_json(path: Path) -> dict | list | None:
    if not path.exists():
        return None
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _style_header(ws, row: int, col_count: int) -> None:
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


def _write_table(ws, headers: list[str], rows: list[list], start_row: int = 1) -> int:
    for col, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=col, value=h)
    _style_header(ws, start_row, len(headers))
    for i, row in enumerate(rows, start_row + 1):
        for col, val in enumerate(row, 1):
            ws.cell(row=i, column=col, value=val)
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    return start_row + len(rows) + 2


def generate(deal_id: str) -> bytes:
    """Build a multi-tab Excel databook and return bytes."""
    deal = deal_store.get_deal(deal_id)
    if deal is None:
        raise DatabookError(f"Deal {deal_id} not found")

    processed = file_store.get_processed_dir(deal_id)
    qoe = _load_json(processed / "qoe_report.json")
    if not qoe:
        raise DatabookError("QoE report not found. Run the full pipeline before exporting databook.")

    wb = Workbook()
    wb.remove(wb.active)

    # Cover
    cover = wb.create_sheet("Cover")
    cover["A1"] = "FDD Databook"
    cover["A2"] = f"Company: {deal['company_name']}"
    cover["A3"] = f"Deal: {deal['deal_name']}"
    cover["A4"] = f"Currency: {deal['currency']}"
    cover["A5"] = f"Generated: {datetime.now(UTC).isoformat()}"

    # QoE Waterfall
    wf = wb.create_sheet("QoE Waterfall")
    wf_rows = []
    for item in qoe.get("waterfall", []):
        wf_rows.append([item.get("label"), item.get("amount"), item.get("type")])
    next_row = _write_table(wf, ["Label", "Amount", "Type"], wf_rows)

    # Normalized EBITDA formula row
    amounts = [Decimal(str(r[1])) for r in wf_rows if r[1] is not None]
    if amounts:
        wf.cell(row=next_row, column=1, value="Check: Waterfall terminal")
        wf.cell(row=next_row, column=2, value=str(sum(amounts)))

    # Adjustment Ledger
    adj_sheet = wb.create_sheet("Adjustment Ledger")
    adj_rows = []
    for adj in qoe.get("adjustments", []):
        adj_rows.append([
            adj.get("adjustment_id"),
            adj.get("label"),
            adj.get("category"),
            adj.get("direction"),
            adj.get("adjustment_amount"),
            ", ".join(adj.get("source_gl_line_ids", [])),
        ])
    _write_table(
        adj_sheet,
        ["ID", "Label", "Category", "Direction", "Amount", "Source GL Lines"],
        adj_rows,
    )

    # GL Mapping
    mapped = _load_json(processed / "mapped_gl.json")
    map_sheet = wb.create_sheet("GL Mapping")
    if mapped:
        seen: dict[str, list] = {}
        for line in mapped:
            code = line.get("account_code", "")
            if code not in seen:
                seen[code] = [
                    code,
                    line.get("account_description"),
                    line.get("standard_category"),
                    line.get("financial_statement"),
                ]
        _write_table(
            map_sheet,
            ["Account Code", "Description", "Category", "Statement"],
            list(seen.values()),
        )

    # Financials tabs
    for fname, title in [
        ("financials_pnl.json", "P&L"),
        ("financials_bs.json", "Balance Sheet"),
        ("financials_cf.json", "Cash Flow"),
    ]:
        data = _load_json(processed / fname)
        if not data:
            continue
        sheet = wb.create_sheet(title)
        rows = data.get("rows", [])
        if rows:
            headers = list(rows[0].keys()) if rows else []
            table_rows = [[r.get(h) for h in headers] for r in rows]
            _write_table(sheet, headers, table_rows)

    # AR/AP Aging
    for fname, title in [("ar_aging.json", "AR Aging"), ("ap_aging.json", "AP Aging")]:
        data = _load_json(processed / fname)
        if not data:
            continue
        sheet = wb.create_sheet(title)
        summaries = data.get("summaries", [])
        aging_rows = [
            [
                s.get("period"),
                str(s.get("bucket_0_30")),
                str(s.get("bucket_31_60")),
                str(s.get("bucket_61_90")),
                str(s.get("bucket_90_plus")),
                str(s.get("total")),
            ]
            for s in summaries
        ]
        _write_table(
            sheet,
            ["Period", "0-30", "31-60", "61-90", "90+", "Total"],
            aging_rows,
        )

    # Cross-doc tie-outs
    cross = _load_json(processed / "cross_document_validation.json")
    if cross:
        tie_sheet = wb.create_sheet("Tie-outs")
        tie_rows = [
            [
                t.get("name"),
                str(t.get("expected")),
                str(t.get("observed")),
                str(t.get("difference")),
                t.get("variance_pct"),
                t.get("status"),
            ]
            for t in cross.get("tie_outs", [])
        ]
        _write_table(
            tie_sheet,
            ["Tie-out", "Expected (GL)", "Observed (Doc)", "Diff", "Variance %", "Status"],
            tie_rows,
        )

    # IRL
    irl_sheet = wb.create_sheet("IRL")
    irl_rows = _build_irl_rows(deal_id, processed, cross)
    _write_table(
        irl_sheet,
        ["Request", "Severity", "Owner", "Source", "Blocking"],
        irl_rows,
    )

    buf = BytesIO()
    wb.save(buf)
    logger.info("Databook generated for deal %s (%d sheets)", deal_id, len(wb.sheetnames))
    return buf.getvalue()


def _build_irl_rows(deal_id: str, processed: Path, cross: dict | None) -> list[list]:
    rows: list[list] = []

    redflags = _load_json(processed / "redflag_report.json")
    if redflags:
        for flag in redflags.get("flags", []):
            for q in flag.get("diligence_questions", []):
                rows.append([q, flag.get("severity", "Medium"), "Management", "Red Flag", "No"])

    if cross:
        for tie in cross.get("tie_outs", []):
            if tie.get("status") in ("Warn", "Fail"):
                rows.append([
                    f"Reconcile {tie.get('name')} variance of {tie.get('difference')}",
                    "Medium",
                    "Controller",
                    "Cross-doc validation",
                    "Yes",
                ])

    inventory = _load_json(processed / "document_inventory.json")
    if inventory:
        for missing in inventory.get("missing_recommended", []):
            rows.append([
                f"Provide {missing.replace('_', ' ')} document",
                "Medium",
                "Data Room Owner",
                "Document inventory",
                "No",
            ])

    if not rows:
        rows.append(["No open diligence requests at this time.", "Informational", "N/A", "System", "No"])

    return rows
