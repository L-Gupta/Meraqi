"""Databook generator tests."""

import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from app.config import settings
from app.pipeline.databook.generator import DatabookError, generate
from app.storage import deal_store, file_store

FIXTURES = Path(__file__).parent.parent / "fixtures"


@pytest.fixture
def deal_with_qoe(tmp_path):
    deal = deal_store.create_deal("Databook Co", "Export Test", "USD")
    deal_id = deal["deal_id"]
    processed = file_store.get_processed_dir(deal_id)

    qoe_report = {
        "deal_id": deal_id,
        "waterfall": [
            {"label": "Reported EBITDA", "amount": "5000000", "type": "base"},
            {"label": "Owner Comp Add-back", "amount": "200000", "type": "addback"},
            {"label": "Adjusted EBITDA", "amount": "5200000", "type": "result"},
        ],
        "adjustments": [
            {
                "adjustment_id": "ADJ-001",
                "label": "Owner Comp Normalisation",
                "category": "Owner/Related Party",
                "direction": "add_back",
                "adjustment_amount": "200000",
                "source_gl_line_ids": ["GL-ABC123"],
            }
        ],
    }
    with open(processed / "qoe_report.json", "w", encoding="utf-8") as f:
        json.dump(qoe_report, f)

    return deal_id


class TestDatabookGenerator:
    def test_raises_without_qoe(self):
        deal = deal_store.create_deal("No QoE", "Fail Test", "USD")
        with pytest.raises(DatabookError, match="QoE report not found"):
            generate(deal["deal_id"])

    def test_generates_valid_xlsx(self, deal_with_qoe):
        content = generate(deal_with_qoe)
        assert content[:2] == b"PK"  # ZIP/XLSX magic

        out = settings.processed_dir / deal_with_qoe / "test_databook.xlsx"
        out.write_bytes(content)
        wb = load_workbook(out)
        assert "Cover" in wb.sheetnames
        assert "QoE Waterfall" in wb.sheetnames
        assert "IRL" in wb.sheetnames

    def test_api_export(self, deal_with_qoe):
        from fastapi.testclient import TestClient

        from app.main import app
        client = TestClient(app)
        resp = client.post(f"/api/v1/deals/{deal_with_qoe}/databook/export")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]
